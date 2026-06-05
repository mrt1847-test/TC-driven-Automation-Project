from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from slugify import slugify
from sqlmodel import Session, select

from worker.core.config import new_id
from worker.models.db import TestCase, TestCaseStatus
from worker.models.schemas import ExcelColumnMapping, ExcelImportRequest, ExcelPreviewRequest, NormalizedTestCase, SourceLocation, TestStep


DEFAULT_MAPPING = ExcelColumnMapping()


def _split_import_lines(raw_text: str) -> list[str]:
    return [line.strip() for line in re.split(r"[\n;]+", raw_text or "") if line.strip()]


def _parse_steps(raw_step: str, raw_expected: str) -> list[TestStep]:
    step_lines = _split_import_lines(raw_step)
    expected_lines = _split_import_lines(raw_expected)
    if not step_lines and raw_step:
        step_lines = [raw_step.strip()]

    # A single expected value with multiple steps is treated as the TC-level
    # expected result only, not as the first step's per-step expectation.
    if len(expected_lines) == 1 and len(step_lines) > 1:
        per_step_expected: list[str | None] = [None] * len(step_lines)
    else:
        per_step_expected = [
            expected_lines[idx - 1] if idx - 1 < len(expected_lines) else None
            for idx in range(1, len(step_lines) + 1)
        ]

    return [
        TestStep(index=idx, action=action, expected=per_step_expected[idx - 1])
        for idx, action in enumerate(step_lines, start=1)
    ]


def _generate_automation_key(title: str, case_id: str, existing: set[str]) -> str:
    base = slugify(case_id or title, separator="_").lower() or "case"
    candidate = base
    counter = 1
    while candidate in existing:
        candidate = f"{base}_{counter:03d}"
        counter += 1
    return candidate


def _load_workbook_rows(file_path: str, sheet_name: str | None, mapping: ExcelColumnMapping) -> tuple[list[str], list[dict[str, Any]]]:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h or "").strip() for h in rows[0]]
    data_rows = []
    for row_index, row in enumerate(rows[1:], start=2):
        row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        row_dict["_row_index"] = row_index
        data_rows.append(row_dict)
    return headers, data_rows


def preview_excel(request: ExcelPreviewRequest) -> dict[str, Any]:
    mapping = request.column_mapping or DEFAULT_MAPPING
    headers, rows = _load_workbook_rows(request.file_path, request.sheet_name, mapping)
    preview = []
    for row in rows[:50]:
        preview.append({
            "rowIndex": row["_row_index"],
            "caseId": row.get(mapping.case_id),
            "title": row.get(mapping.title),
            "automationKey": row.get(mapping.automation_key),
            "step": row.get(mapping.step),
            "expected": row.get(mapping.expected),
            "startUrl": row.get(mapping.start_url),
        })
    return {"headers": headers, "preview": preview, "totalRows": len(rows)}


def import_excel(session: Session, project_id: str, request: ExcelImportRequest) -> list[NormalizedTestCase]:
    mapping = request.column_mapping or DEFAULT_MAPPING
    _, rows = _load_workbook_rows(request.file_path, request.sheet_name, mapping)
    existing_keys = {tc.automation_key for tc in session.exec(select(TestCase).where(TestCase.project_id == project_id)).all()}
    imported: list[NormalizedTestCase] = []

    for row in rows:
        row_index = row["_row_index"]
        if request.selected_rows and row_index not in request.selected_rows:
            continue
        case_id = str(row.get(mapping.case_id) or f"ROW-{row_index}")
        title = str(row.get(mapping.title) or case_id)
        automation_key = str(row.get(mapping.automation_key) or "").strip()
        if not automation_key:
            automation_key = _generate_automation_key(title, case_id, existing_keys)
        existing_keys.add(automation_key)

        preconditions = []
        pre_raw = row.get(mapping.precondition)
        if pre_raw:
            preconditions = [p.strip() for p in str(pre_raw).split("\n") if p.strip()]

        normalized = NormalizedTestCase(
            id=new_id("tc"),
            source_type="excel",
            source_id=case_id,
            source_location=SourceLocation(
                file_path=request.file_path,
                sheet_name=request.sheet_name,
                row_index=row_index,
            ),
            title=title,
            preconditions=preconditions,
            steps=_parse_steps(str(row.get(mapping.step) or ""), str(row.get(mapping.expected) or "")),
            expected_result=str(row.get(mapping.expected) or "") or None,
            automation_key=automation_key,
            priority=str(row.get(mapping.priority) or "") or None,
            start_url=str(row.get(mapping.start_url) or "") or None,
        )

        db_case = TestCase(
            id=normalized.id,
            project_id=project_id,
            source_type=normalized.source_type,
            source_case_id=normalized.source_id,
            source_location_json=json.dumps(normalized.source_location.model_dump() if normalized.source_location else {}),
            title=normalized.title,
            preconditions_json=json.dumps(normalized.preconditions),
            steps_json=json.dumps([s.model_dump() for s in normalized.steps]),
            expected_result=normalized.expected_result,
            automation_key=normalized.automation_key,
            priority=normalized.priority,
            start_url=normalized.start_url,
            status=TestCaseStatus.imported.value,
        )
        session.add(db_case)
        imported.append(normalized)

    session.commit()
    return imported


def case_to_normalized(case: TestCase) -> NormalizedTestCase:
    return NormalizedTestCase(
        id=case.id,
        source_type=case.source_type,
        source_id=case.source_case_id,
        source_location=json.loads(case.source_location_json) if case.source_location_json else None,
        title=case.title,
        preconditions=json.loads(case.preconditions_json or "[]"),
        steps=[TestStep.model_validate(s) for s in json.loads(case.steps_json or "[]")],
        expected_result=case.expected_result,
        automation_key=case.automation_key,
        priority=case.priority,
        start_url=case.start_url,
        status=case.status,
    )
