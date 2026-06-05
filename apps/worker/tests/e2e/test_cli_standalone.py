"""B2-07: generated-template runner.cli works without the Worker."""

from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[4]
TEMPLATE = ROOT / "packages" / "generated-template"


def run_cli(project: Path, *args: str, extra_env: dict[str, str] | None = None) -> subprocess.CompletedProcess[str]:
    result = subprocess.run(
        [sys.executable, "-m", "runner.cli", *args],
        cwd=project,
        env={**os.environ, **(extra_env or {}), "PYTHONPATH": str(project)},
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0, (
        f"runner.cli {' '.join(args)} failed\n"
        f"stdout:\n{result.stdout}\n"
        f"stderr:\n{result.stderr}"
    )
    return result


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def prepare_generated_project(tmp_path: Path) -> Path:
    project = tmp_path / "generated-project"
    shutil.copytree(
        TEMPLATE,
        project,
        ignore=shutil.ignore_patterns("__pycache__", ".pytest_cache"),
    )

    write_text(
        project / "mappings" / "cases.yaml",
        """
cases:
  - automationKey: cli_pass
    sourceType: excel
    sourceCaseId: TC-PASS
    title: CLI passing case
    testFile: tests/test_cli_pass.py
    testFunction: test_cli_pass
    tags:
      - standalone
    resultTargets:
      excel:
        file: imported_cases.xlsx
        sheet: TestCases
        row: 2
  - automationKey: cli_fail
    sourceType: excel
    sourceCaseId: TC-FAIL
    title: CLI failing case
    testFile: tests/test_cli_fail.py
    testFunction: test_cli_fail
    tags:
      - standalone
    resultTargets:
      excel:
        file: imported_cases.xlsx
        sheet: TestCases
        row: 3
""".lstrip(),
    )
    write_text(
        project / "tests" / "test_cli_pass.py",
        """
import os


def test_cli_pass(pytestconfig):
    assert os.environ["TC_ENV"] == "stg"
    assert os.environ["TC_RUN_ID"].startswith("cli-")
    assert os.environ["TC_BROWSER"] == "chromium"
    assert os.environ["TC_HEADLESS"] == "true"
    browser_option = pytestconfig.getoption("browser")
    assert browser_option == "chromium" or browser_option == ["chromium"]
""".lstrip(),
    )
    write_text(
        project / "tests" / "test_cli_fail.py",
        """
def test_cli_fail():
    assert False, "intentional standalone failure for rerun-failed coverage"
""".lstrip(),
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TestCases"
    sheet.append(
        [
            "Case ID",
            "Title",
            "Automation Result",
            "Automation Run ID",
            "Automation Executed At",
            "Automation Comment",
        ]
    )
    sheet.append(["TC-PASS", "CLI passing case", None, None, None, None])
    sheet.append(["TC-FAIL", "CLI failing case", None, None, None, None])
    workbook.save(project / "imported_cases.xlsx")
    return project


def read_results(project: Path, run_id: str) -> dict:
    path = project / "artifacts" / "runs" / run_id / "results.json"
    assert path.exists(), f"Missing results.json for {run_id}"
    return json.loads(path.read_text(encoding="utf-8"))


def test_generated_template_cli_standalone_e2e(tmp_path: Path) -> None:
    project = prepare_generated_project(tmp_path)

    listed = run_cli(project, "list-cases")
    assert "cli_pass\tCLI passing case" in listed.stdout
    assert "cli_fail\tCLI failing case" in listed.stdout

    run_cli(
        project,
        "run",
        "--env",
        "stg",
        "--browser",
        "chromium",
        "--case-key",
        "cli_pass",
        "--run-id",
        "cli-pass-run",
    )
    pass_results = read_results(project, "cli-pass-run")
    assert pass_results["summary"] == {"total": 1, "passed": 1, "failed": 0, "skipped": 0}
    assert pass_results["pytest"]["returnCode"] == 0
    assert pass_results["pytest"]["stdoutPath"] == "artifacts/runs/cli-pass-run/stdout.log"
    assert (project / pass_results["pytest"]["stdoutPath"]).exists()
    assert (project / pass_results["pytest"]["stderrPath"]).exists()
    assert pass_results["cases"][0]["automationKey"] == "cli_pass"
    assert pass_results["cases"][0]["artifacts"] == {"screenshot": None, "trace": None, "video": None}

    run_cli(
        project,
        "run",
        "--env",
        "stg",
        "--browser",
        "chromium",
        "--all",
        "--run-id",
        "cli-all-run",
    )
    all_results = read_results(project, "cli-all-run")
    assert all_results["pytest"]["returnCode"] != 0
    statuses = {case["automationKey"]: case["status"] for case in all_results["cases"]}
    assert statuses == {"cli_pass": "passed", "cli_fail": "failed"}
    failed_case = next(case for case in all_results["cases"] if case["automationKey"] == "cli_fail")
    assert "intentional standalone failure" in failed_case["error"]

    run_cli(
        project,
        "rerun-failed",
        "--from-run-id",
        "cli-all-run",
        "--run-id",
        "cli-rerun",
    )
    rerun_results = read_results(project, "cli-rerun")
    assert [case["automationKey"] for case in rerun_results["cases"]] == ["cli_fail"]
    assert rerun_results["summary"]["failed"] == 1

    exported = run_cli(project, "export", "--run-id", "cli-all-run", "--target", "excel")
    assert "Export completed" in exported.stdout

    workbook = load_workbook(project / "imported_cases.xlsx")
    sheet = workbook.active
    assert sheet["C2"].value == "passed"
    assert sheet["D2"].value == "cli-all-run"
    assert sheet["C3"].value == "failed"
    assert sheet["D3"].value == "cli-all-run"


def test_generated_template_runner_redacts_secret_values_from_artifacts(tmp_path: Path) -> None:
    project = tmp_path / "generated-project"
    shutil.copytree(
        TEMPLATE,
        project,
        ignore=shutil.ignore_patterns("__pycache__", ".pytest_cache"),
    )
    write_text(
        project / "mappings" / "cases.yaml",
        """
cases:
  - automationKey: cli_secret
    sourceType: excel
    sourceCaseId: TC-SECRET
    title: CLI secret redaction case
    testFile: tests/test_cli_secret.py
    testFunction: test_cli_secret
""".lstrip(),
    )
    write_text(
        project / "tests" / "test_cli_secret.py",
        """
import os
import sys


def test_cli_secret():
    secret = os.environ["OPENAI_API_KEY"]
    print(f"stdout secret={secret}")
    print(f"stderr secret={secret}", file=sys.stderr)
    assert False, f"failure leaked {secret}"
""".lstrip(),
    )

    secret = "value-visible-only-via-env-123456789"
    run_cli(
        project,
        "run",
        "--env",
        "stg",
        "--browser",
        "chromium",
        "--case-key",
        "cli_secret",
        "--run-id",
        "cli-secret-run",
        extra_env={"OPENAI_API_KEY": secret},
    )

    run_dir = project / "artifacts" / "runs" / "cli-secret-run"
    stdout_text = (run_dir / "stdout.log").read_text(encoding="utf-8")
    stderr_text = (run_dir / "stderr.log").read_text(encoding="utf-8")
    results_text = (run_dir / "results.json").read_text(encoding="utf-8")

    assert secret not in stdout_text
    assert secret not in stderr_text
    assert secret not in results_text
    assert "***MASKED***" in stdout_text
    assert "***MASKED***" in results_text
