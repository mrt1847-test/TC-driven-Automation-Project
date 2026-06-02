"""H-04: MVP 4 gate - remaining result write-back targets."""
from __future__ import annotations

from pathlib import Path
import shutil
import time

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlmodel import Session, select

from worker.models.db import ExportLog

ROOT = Path(__file__).resolve().parents[4]
EXCEL_FIXTURE = ROOT / "fixtures" / "sample_cases.xlsx"


def _wait_for_run(client: TestClient, project_id: str, case_id: str, timeout_s: float = 5.0) -> dict:
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        runs = client.get(f"/projects/{project_id}/webwright-runs").json()
        for run in runs:
            if run.get("test_case_id") == case_id and run.get("status") in {"completed", "failed", "cancelled"}:
                return run
        time.sleep(0.05)
    pytest.fail("Timed out waiting for Webwright run to finish")


def _wait_for_execution(client: TestClient, project_id: str, timeout_s: float = 8.0) -> dict:
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        executions = client.get(f"/projects/{project_id}/executions").json()
        for execution in executions:
            if execution.get("status") in {"completed", "failed", "cancelled"}:
                return execution
        time.sleep(0.05)
    pytest.fail("Timed out waiting for execution to finish")


def _import_excel_copy(client: TestClient, project_id: str, tmp_path: Path) -> dict:
    if not EXCEL_FIXTURE.exists():
        pytest.skip(f"Missing Excel fixture: {EXCEL_FIXTURE}")
    source = tmp_path / "mvp4_cases.xlsx"
    shutil.copy2(EXCEL_FIXTURE, source)
    response = client.post(f"/projects/{project_id}/cases/import/excel", json={"file_path": str(source)})
    assert response.status_code == 200
    cases = response.json()
    assert cases
    return cases[0]


def _save_reviewed_mappings(client: TestClient, project_id: str, case_id: str) -> None:
    mappings = client.get(f"/projects/{project_id}/cases/{case_id}/mappings").json()
    assert mappings
    reviewed = []
    for index, mapping in enumerate(mappings, start=1):
        reviewed.append({
            **mapping,
            "normalized_step_id": mapping.get("normalized_step_id") or f"flow_{index:03d}",
            "normalized_step_name": f"mvp4_step_{index}",
            "pom_method_name": f"perform_mvp4_step_{index}",
            "status": "mapped",
        })
    save = client.put(f"/projects/{project_id}/cases/{case_id}/mappings", json={"mappings": reviewed})
    assert save.status_code == 200


def _write_fake_pytest_fixture(client: TestClient, project_id: str) -> None:
    content = "\n".join([
        "import pytest",
        "",
        "",
        "def pytest_addoption(parser):",
        "    parser.addoption('--browser', action='store', default='chromium')",
        "    parser.addoption('--headed', action='store', default='false')",
        "",
        "",
        "class FakeLocator:",
        "    def click(self):",
        "        return None",
        "    def fill(self, value):",
        "        return None",
        "",
        "",
        "class FakePage:",
        "    def goto(self, url):",
        "        return None",
        "    def locator(self, selector):",
        "        return FakeLocator()",
        "    def get_by_role(self, *args, **kwargs):",
        "        return FakeLocator()",
        "",
        "",
        "@pytest.fixture",
        "def page():",
        "    return FakePage()",
        "",
    ])
    response = client.put(
        f"/projects/{project_id}/generated-files/content",
        json={"path": "conftest.py", "content": content},
    )
    assert response.status_code == 200


def _prepare_execution(client: TestClient, project_id: str, case: dict) -> dict:
    case_id = case["id"]
    queued = client.post(f"/projects/{project_id}/webwright-runs", json={"caseIds": [case_id]})
    assert queued.status_code == 200
    run = _wait_for_run(client, project_id, case_id)
    assert run["status"] == "completed"

    _save_reviewed_mappings(client, project_id, case_id)
    generated = client.post(f"/projects/{project_id}/generate", json={"caseIds": [case_id]})
    assert generated.status_code == 200
    assert Path(generated.json()["generatedProjectPath"]).exists()
    _write_fake_pytest_fixture(client, project_id)

    execution_queued = client.post(
        f"/projects/{project_id}/executions",
        json={
            "env": "stg",
            "browser": "chromium",
            "headed": False,
            "target_type": "case",
            "automation_key": case["automation_key"],
            "result_target": "local",
        },
    )
    assert execution_queued.status_code == 200
    execution = _wait_for_execution(client, project_id)
    assert execution["result_path"]
    assert Path(execution["result_path"]).exists()
    return execution


def _assert_preview_update(payload: dict, case: dict, execution: dict) -> dict:
    assert payload["preview"] is True
    updates = payload["updates"]
    assert len(updates) == 1
    update = updates[0]
    assert update["automationKey"] == case["automation_key"]
    assert update["sourceCaseId"] == case["source_id"]
    assert update["runId"] == execution["run_id"]
    return update


def _workbook_value(path: str, row: int, column_name: str):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    return sheet.cell(row=row, column=headers.index(column_name) + 1).value


def test_mvp4_remaining_write_back_targets_gate(
    client: TestClient,
    project_id: str,
    tmp_path: Path,
) -> None:
    case = _import_excel_copy(client, project_id, tmp_path)
    execution = _prepare_execution(client, project_id, case)

    detail = client.get(f"/projects/{project_id}/executions/{execution['id']}").json()
    result = next(item for item in detail["results"] if item["automation_key"] == case["automation_key"])
    assert result["source_case_id"] == case["source_id"]

    testrail_preview = client.post(
        f"/projects/{project_id}/executions/{execution['id']}/export/testrail",
        json={"preview": True},
    )
    assert testrail_preview.status_code == 200
    _assert_preview_update(testrail_preview.json(), case, execution)

    testrail_export = client.post(
        f"/projects/{project_id}/executions/{execution['id']}/export/testrail",
        json={"preview": False},
    )
    assert testrail_export.status_code == 200
    assert testrail_export.json()["updated"] == 1
    assert testrail_export.json()["target"] == "testrail"

    google_preview = client.post(
        f"/projects/{project_id}/executions/{execution['id']}/export/google-sheets",
        json={"preview": True},
    )
    assert google_preview.status_code == 200
    _assert_preview_update(google_preview.json(), case, execution)

    google_export = client.post(
        f"/projects/{project_id}/executions/{execution['id']}/export/google-sheets",
        json={"preview": False},
    )
    assert google_export.status_code == 200
    assert google_export.json()["updated"] == 1
    assert google_export.json()["target"] == "google-sheets"

    excel_preview = client.post(
        f"/projects/{project_id}/executions/{execution['id']}/export/excel",
        json={"preview": True},
    )
    assert excel_preview.status_code == 200
    excel_update = _assert_preview_update(excel_preview.json(), case, execution)

    excel_export = client.post(
        f"/projects/{project_id}/executions/{execution['id']}/export/excel",
        json={"preview": False},
    )
    assert excel_export.status_code == 200
    assert excel_export.json()["updated"] == 1
    assert excel_export.json()["failed"] == []
    assert _workbook_value(excel_update["file"], excel_update["row"], "Automation Run ID") == execution["run_id"]

    cases_yaml = client.get(
        f"/projects/{project_id}/generated-files/content",
        params={"path": "mappings/cases.yaml"},
    ).json()["content"]
    missing_file = tmp_path / "missing_source.xlsx"
    broken_yaml = cases_yaml.replace(excel_update["file"].replace("\\", "\\\\"), str(missing_file).replace("\\", "\\\\"))
    if broken_yaml == cases_yaml:
        broken_yaml = cases_yaml.replace(excel_update["file"], str(missing_file))
    patch_mapping = client.put(
        f"/projects/{project_id}/generated-files/content",
        json={"path": "mappings/cases.yaml", "content": broken_yaml},
    )
    assert patch_mapping.status_code == 200

    failed_excel_export = client.post(
        f"/projects/{project_id}/executions/{execution['id']}/export/excel",
        json={"preview": False},
    )
    assert failed_excel_export.status_code == 200
    failed_body = failed_excel_export.json()
    assert failed_body["updated"] == 0
    assert len(failed_body["failed"]) == 1
    assert failed_body["failed"][0]["sourceCaseId"] == case["source_id"]
    assert failed_body["failed"][0]["error"] == "source file not found"
    assert Path(execution["result_path"]).exists()
    assert _workbook_value(excel_update["file"], excel_update["row"], "Automation Run ID") == execution["run_id"]

    import worker.core.database as database

    with Session(database.engine) as session:
        logs = session.exec(select(ExportLog).where(ExportLog.execution_run_id == execution["id"])).all()

    logs_by_target = {}
    for log in logs:
        logs_by_target.setdefault(log.target, []).append(log)

    assert logs_by_target["testrail"][0].status == "success"
    assert logs_by_target["google-sheets"][0].status == "success"
    assert any(log.status == "success" for log in logs_by_target["excel"])
    assert any(log.status == "failed" and case["source_id"] in (log.message or "") for log in logs_by_target["excel"])
