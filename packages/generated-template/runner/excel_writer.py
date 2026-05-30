from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from runner.mapping_loader import load_cases, project_root


def write(results_path: Path) -> None:
    data = __import__("json").loads(results_path.read_text(encoding="utf-8"))
    by_key = {c["automationKey"]: c for c in load_cases()}
    for case in data.get("cases", []):
        meta = by_key.get(case.get("automationKey"), {})
        excel = (meta.get("resultTargets") or {}).get("excel") or {}
        file_path = excel.get("file")
        if not file_path:
            continue
        src = Path(file_path)
        if not src.is_absolute():
            src = project_root() / src
        if not src.exists():
            continue
        backup = src.with_suffix(src.suffix + f".backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}")
        shutil.copy2(src, backup)
        wb = load_workbook(src)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        row = excel.get("row", 2)
        updates = {
            "Automation Result": case.get("status"),
            "Automation Run ID": data.get("runId"),
            "Automation Executed At": datetime.utcnow().isoformat(),
            "Automation Comment": case.get("error") or "",
        }
        for col, val in updates.items():
            if col in headers:
                ws.cell(row=row, column=headers.index(col) + 1, value=val)
        wb.save(src)
