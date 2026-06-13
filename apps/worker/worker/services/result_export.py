from __future__ import annotations

import json
import math
import shutil
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import httpx
from openpyxl import load_workbook
from typing import Any

from sqlmodel import Session, select

from worker.core.config import MASK, load_settings, mask_secret_data, mask_secrets, new_id
from worker.models.db import ExportLog, ExecutionResult, ExecutionRun
from worker.services.automation_keys import duplicate_active_automation_keys
from worker.services.integrations.google_sheets import (
    GoogleSheetsConnectorError,
    SHEETS_SCOPE,
    _google_access_token,
    _mask_with_credential,
)


class ExportValidationError(ValueError):
    """Raised when an export would write back to an unsafe target mapping."""


TESTRAIL_STATUS_IDS = {
    "passed": 1,
    "pass": 1,
    "success": 1,
    "blocked": 2,
    "untested": 3,
    "retest": 4,
    "failed": 5,
    "fail": 5,
    "error": 5,
    "cancelled": 2,
    "canceled": 2,
    "skipped": 2,
}

GOOGLE_SHEETS_RESULT_HEADERS = {
    "status": "Automation Result",
    "run_id": "Automation Run ID",
    "executed_at": "Automation Executed At",
    "comment": "Automation Comment",
}


def _load_results(execution: ExecutionRun) -> dict[str, Any]:
    if not execution.result_path or not Path(execution.result_path).exists():
        raise FileNotFoundError("results.json not found")
    return json.loads(Path(execution.result_path).read_text(encoding="utf-8"))


def _result_updates(execution: ExecutionRun, results: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    results = results if results is not None else _load_results(execution)
    return [
        {
            "automationKey": case.get("automationKey"),
            "sourceType": case.get("sourceType"),
            "sourceCaseId": case.get("sourceCaseId"),
            "title": case.get("title"),
            "status": case.get("status"),
            "durationMs": case.get("durationMs"),
            "runId": execution.run_id,
            "comment": case.get("error") or "Automation passed",
            "artifacts": case.get("artifacts") or {},
        }
        for case in results.get("cases", [])
    ]


def _generated_project_root(execution: ExecutionRun) -> Path:
    if not execution.result_path:
        raise FileNotFoundError("results.json not found")
    try:
        return Path(execution.result_path).parents[3]
    except IndexError as exc:
        raise ExportValidationError("results.json path is not inside a generated project run directory") from exc


def _load_mapping_entries(execution: ExecutionRun) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    mapping_path = _generated_project_root(execution) / "mappings" / "cases.yaml"
    if not mapping_path.exists():
        return [], [{
            "kind": "mapping_file_missing",
            "message": "Generated mappings/cases.yaml was not found",
            "path": str(mapping_path),
        }]

    try:
        import yaml

        mapping_data = yaml.safe_load(mapping_path.read_text(encoding="utf-8")) or {}
    except Exception as exc:
        return [], [{
            "kind": "mapping_load_failed",
            "message": f"Generated mappings/cases.yaml could not be loaded: {exc}",
            "path": str(mapping_path),
        }]

    cases = mapping_data.get("cases")
    if not isinstance(cases, list):
        return [], [{
            "kind": "mapping_cases_invalid",
            "message": "Generated mappings/cases.yaml must contain a cases list",
            "path": str(mapping_path),
        }]
    return cases, []


def _group_by_key(items: list[Any], key_getter) -> dict[str | None, list[Any]]:
    grouped: dict[str | None, list[Any]] = {}
    for item in items:
        grouped.setdefault(key_getter(item), []).append(item)
    return grouped


def _issue(kind: str, message: str, automation_key: str | None = None, **extra: Any) -> dict[str, Any]:
    item = {"kind": kind, "message": message}
    if automation_key is not None:
        item["automationKey"] = automation_key
    item.update(extra)
    return item


def _compare_identity(
    issues: list[dict[str, Any]],
    *,
    automation_key: str,
    left_name: str,
    right_name: str,
    left: dict[str, Any],
    right: dict[str, Any],
) -> None:
    left_source_type = left.get("sourceType")
    right_source_type = right.get("sourceType")
    if left_source_type != right_source_type:
        issues.append(_issue(
            "source_type_mismatch",
            f"{left_name} sourceType does not match {right_name}",
            automation_key,
            resultValue=left_source_type,
            expectedValue=right_source_type,
        ))

    left_source_case_id = left.get("sourceCaseId")
    right_source_case_id = right.get("sourceCaseId")
    if left_source_case_id != right_source_case_id:
        issues.append(_issue(
            "source_case_id_mismatch",
            f"{left_name} sourceCaseId does not match {right_name}",
            automation_key,
            resultValue=left_source_case_id,
            expectedValue=right_source_case_id,
        ))


def _db_result_identity(result: ExecutionResult) -> dict[str, Any]:
    return {
        "sourceType": result.source_type,
        "sourceCaseId": result.source_case_id,
    }


def _mapping_identity(mapping: dict[str, Any]) -> dict[str, Any]:
    return {
        "sourceType": mapping.get("sourceType"),
        "sourceCaseId": mapping.get("sourceCaseId"),
    }


def _validate_export(
    session: Session,
    execution: ExecutionRun,
    updates: list[dict[str, Any]],
    mapping_entries: list[dict[str, Any]],
    preload_issues: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    issues = list(preload_issues or [])

    mapping_groups = _group_by_key(mapping_entries, lambda item: item.get("automationKey"))
    update_groups = _group_by_key(updates, lambda item: item.get("automationKey"))
    db_results = session.exec(
        select(ExecutionResult).where(ExecutionResult.execution_run_id == execution.id)
    ).all()
    db_groups = _group_by_key(db_results, lambda item: item.automation_key)
    active_case_duplicates = duplicate_active_automation_keys(session, execution.project_id)

    for key, case_ids in active_case_duplicates.items():
        issues.append(_issue(
            "ambiguous_active_case",
            "Project contains duplicate active test case automation_key values",
            key,
            caseIds=case_ids,
        ))

    for key, rows in mapping_groups.items():
        if key is None:
            issues.append(_issue("missing_automation_key", "Generated mapping row is missing automationKey"))
        elif len(rows) > 1:
            issues.append(_issue(
                "ambiguous_mapping",
                "Generated mappings/cases.yaml contains duplicate automationKey rows",
                key,
            ))

    for key, rows in db_groups.items():
        if not key:
            issues.append(_issue("missing_automation_key", "ExecutionResult row is missing automation_key"))
        elif len(rows) > 1:
            issues.append(_issue(
                "ambiguous_execution_result",
                "ExecutionResult rows contain duplicate automation_key values",
                key,
            ))

    for key, rows in update_groups.items():
        if key and len(rows) > 1:
            issues.append(_issue(
                "ambiguous_result_update",
                "Result rows contain duplicate automationKey values",
                key,
            ))

    for update in updates:
        automation_key = update.get("automationKey")
        if not automation_key:
            issues.append(_issue("missing_automation_key", "Result row is missing automationKey"))
            continue

        mapping_matches = mapping_groups.get(automation_key, [])
        if not mapping_matches:
            issues.append(_issue(
                "missing_mapping",
                "No generated mapping row exists for result automationKey",
                automation_key,
            ))
        elif len(mapping_matches) == 1:
            _compare_identity(
                issues,
                automation_key=automation_key,
                left_name="Result row",
                right_name="generated mapping",
                left=update,
                right=_mapping_identity(mapping_matches[0]),
            )

        db_matches = db_groups.get(automation_key, [])
        if not db_matches:
            issues.append(_issue(
                "missing_execution_result",
                "No ExecutionResult row exists for result automationKey",
                automation_key,
            ))
        elif len(db_matches) == 1:
            _compare_identity(
                issues,
                automation_key=automation_key,
                left_name="Result row",
                right_name="ExecutionResult",
                left=update,
                right=_db_result_identity(db_matches[0]),
            )

    return {
        "ok": not issues,
        "checked": len(updates),
        "issues": issues,
    }


def _raise_if_invalid(validation: dict[str, Any]) -> None:
    if validation.get("ok"):
        return
    kinds = sorted({issue.get("kind", "unknown") for issue in validation.get("issues", [])})
    raise ExportValidationError(f"Export validation failed: {', '.join(kinds)}")


def _merge_validation_issues(validation: dict[str, Any], issues: list[dict[str, Any]]) -> dict[str, Any]:
    merged = {
        "ok": validation.get("ok") and not issues,
        "checked": validation.get("checked", 0),
        "issues": [*validation.get("issues", []), *issues],
    }
    return merged


def _log_export(
    session: Session,
    execution: ExecutionRun,
    target: str,
    status: str,
    updates: list[dict[str, Any]],
    extra: dict[str, Any] | None = None,
) -> None:
    log_payload = {
        "runId": execution.run_id,
        "updates": [
            {
                "automationKey": item.get("automationKey"),
                "sourceCaseId": item.get("sourceCaseId"),
                "status": item.get("status"),
            }
            for item in updates
        ],
    }
    if extra:
        log_payload.update(extra)
    session.add(ExportLog(
        id=new_id("exp"),
        execution_run_id=execution.id,
        target=target,
        status=status,
        message=json.dumps(mask_secret_data(log_payload), ensure_ascii=False),
    ))
    session.commit()


def _testrail_export_config(config: dict[str, Any] | None = None) -> dict[str, Any]:
    request_config = config or {}
    settings = load_settings()
    integration = settings.integrations.get("testrail", {})
    enabled = bool(integration.get("enabled"))
    return {
        "enabled": enabled,
        "mock": bool(request_config.get("mock")) or bool(integration.get("mock")) or not enabled,
        "base_url": (
            request_config.get("baseUrl")
            or request_config.get("base_url")
            or integration.get("baseUrl")
            or integration.get("base_url")
            or ""
        ),
        "username": request_config.get("username") or integration.get("username") or "",
        "api_token": request_config.get("apiToken") or request_config.get("api_token") or request_config.get("token") or "",
        "run_id": (
            request_config.get("runId")
            or request_config.get("run_id")
            or request_config.get("testrailRunId")
            or request_config.get("testrail_run_id")
            or integration.get("runId")
            or integration.get("run_id")
            or integration.get("resultRunId")
            or integration.get("result_run_id")
            or ""
        ),
        "status_ids": request_config.get("statusIds") or request_config.get("status_ids") or integration.get("statusIds") or {},
    }


def _testrail_missing_config(config: dict[str, Any]) -> list[str]:
    required = {
        "baseUrl": config.get("base_url"),
        "username": config.get("username"),
        "apiToken": config.get("api_token"),
        "runId": config.get("run_id"),
    }
    return [name for name, value in required.items() if not str(value or "").strip()]


def _testrail_payloads(
    updates: list[dict[str, Any]],
    mapping_entries: list[dict[str, Any]],
    config: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    mapping_by_key = {item.get("automationKey"): item for item in mapping_entries if item.get("automationKey")}
    payloads: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    for item in updates:
        automation_key = item.get("automationKey")
        mapping = mapping_by_key.get(automation_key, {})
        target = (mapping.get("resultTargets") or {}).get("testrail") or {}
        run_id = str(
            target.get("runId")
            or target.get("run_id")
            or target.get("testrailRunId")
            or target.get("testrail_run_id")
            or config.get("run_id")
            or ""
        ).strip()
        case_id = str(
            target.get("caseId")
            or target.get("case_id")
            or target.get("testrailCaseId")
            or target.get("testrail_case_id")
            or item.get("sourceCaseId")
            or ""
        ).strip()
        if not run_id:
            issues.append(_issue("testrail_run_id_missing", "TestRail run ID is required for result write-back", automation_key))
            continue
        if not case_id:
            issues.append(_issue("testrail_case_id_missing", "TestRail case ID is required for result write-back", automation_key))
            continue
        body = {
            "status_id": _testrail_status_id(str(item.get("status") or ""), config),
            "comment": _testrail_comment(item),
        }
        elapsed = _testrail_elapsed(item.get("durationMs"))
        if elapsed:
            body["elapsed"] = elapsed
        payloads.append({
            "automationKey": automation_key,
            "sourceCaseId": item.get("sourceCaseId"),
            "testrailRunId": run_id,
            "testrailCaseId": case_id,
            "endpoint": _testrail_api_url(str(config.get("base_url") or ""), f"add_result_for_case/{run_id}/{case_id}"),
            "body": body,
        })
    return payloads, issues


def _testrail_status_id(status: str, config: dict[str, Any]) -> int:
    custom = config.get("status_ids") if isinstance(config.get("status_ids"), dict) else {}
    normalized = status.strip().lower()
    raw = custom.get(normalized) or custom.get(status) or TESTRAIL_STATUS_IDS.get(normalized, 5)
    try:
        return int(raw)
    except (TypeError, ValueError):
        return 5


def _testrail_comment(item: dict[str, Any]) -> str:
    title = item.get("title") or item.get("automationKey") or "Automation result"
    automation_key = item.get("automationKey") or ""
    status = item.get("status") or "unknown"
    comment = item.get("comment") or "Automation passed"
    run_id = item.get("runId") or ""
    return f"{title}\nAutomation Key: {automation_key}\nStatus: {status}\nRun: {run_id}\n{comment}".strip()


def _testrail_elapsed(duration_ms: Any) -> str | None:
    try:
        value = int(duration_ms)
    except (TypeError, ValueError):
        return None
    if value <= 0:
        return None
    seconds = max(1, math.ceil(value / 1000))
    return f"{seconds}s"


def _testrail_api_url(base_url: str, endpoint: str) -> str:
    return f"{base_url.rstrip('/')}/index.php?/api/v2/{endpoint.lstrip('/')}"


async def _post_testrail_results(payloads: list[dict[str, Any]], config: dict[str, Any]) -> list[dict[str, Any]]:
    username = str(config.get("username") or "")
    api_token = str(config.get("api_token") or "")
    responses: list[dict[str, Any]] = []
    async with httpx.AsyncClient(timeout=30) as client:
        for payload in payloads:
            try:
                response = await client.post(
                    payload["endpoint"],
                    json=payload["body"],
                    auth=(username, api_token),
                    headers={"Accept": "application/json"},
                )
            except httpx.HTTPError as error:
                raise ExportValidationError(
                    _mask_with_token(f"TestRail API request failed: {error}", api_token)
                ) from error
            if response.status_code >= 400:
                raise ExportValidationError(
                    f"TestRail API returned HTTP {response.status_code}: {_testrail_error_detail(response, api_token)}"
                )
            responses.append({
                "automationKey": payload.get("automationKey"),
                "sourceCaseId": payload.get("sourceCaseId"),
                "statusCode": response.status_code,
                "response": _safe_response_json(response),
            })
    return responses


def _safe_response_json(response: httpx.Response) -> Any:
    try:
        return mask_secret_data(response.json())
    except ValueError:
        return mask_secrets(response.text[:300])


def _testrail_error_detail(response: httpx.Response, api_token: str) -> str:
    try:
        payload = response.json()
    except ValueError:
        payload = response.text
    if isinstance(payload, dict):
        detail = payload.get("error") or payload.get("message") or str(payload)
    else:
        detail = str(payload)
    return _mask_with_token(detail, api_token)


def _mask_with_token(message: str, api_token: str) -> str:
    masked = mask_secrets(message, {"TESTRAIL_API_TOKEN": api_token})
    return masked.replace(api_token, MASK) if api_token else masked


def _google_sheets_export_config(config: dict[str, Any] | None = None) -> dict[str, Any]:
    request_config = config or {}
    settings = load_settings()
    integration = settings.integrations.get("googleSheets", {})
    enabled = bool(integration.get("enabled"))
    result_columns = _merged_dict(
        integration.get("resultColumns"),
        integration.get("result_columns"),
        request_config.get("resultColumns"),
        request_config.get("result_columns"),
    )
    return {
        "enabled": enabled,
        "mock": bool(request_config.get("mock")) or bool(integration.get("mock")) or not enabled,
        "spreadsheet_id": (
            request_config.get("spreadsheetId")
            or request_config.get("spreadsheet_id")
            or integration.get("spreadsheetId")
            or integration.get("spreadsheet_id")
            or ""
        ),
        "sheet_name": (
            request_config.get("sheetName")
            or request_config.get("sheet_name")
            or request_config.get("resultSheetName")
            or request_config.get("result_sheet_name")
            or integration.get("sheetName")
            or integration.get("sheet_name")
            or integration.get("resultSheetName")
            or integration.get("result_sheet_name")
            or "Cases"
        ),
        "credential_json": (
            request_config.get("credentialJson")
            or request_config.get("credential_json")
            or request_config.get("serviceAccountJson")
            or request_config.get("service_account_json")
            or request_config.get("accessToken")
            or request_config.get("access_token")
            or ""
        ),
        "header_row": _positive_int(
            request_config.get("headerRow")
            or request_config.get("header_row")
            or integration.get("headerRow")
            or integration.get("header_row"),
            1,
        ),
        "value_input_option": (
            request_config.get("valueInputOption")
            or request_config.get("value_input_option")
            or integration.get("valueInputOption")
            or integration.get("value_input_option")
            or "USER_ENTERED"
        ),
        "headers": {
            "status": _result_header(result_columns, "status", "result", default=GOOGLE_SHEETS_RESULT_HEADERS["status"]),
            "run_id": _result_header(result_columns, "runId", "run_id", default=GOOGLE_SHEETS_RESULT_HEADERS["run_id"]),
            "executed_at": _result_header(
                result_columns,
                "executedAt",
                "executed_at",
                default=GOOGLE_SHEETS_RESULT_HEADERS["executed_at"],
            ),
            "comment": _result_header(result_columns, "comment", default=GOOGLE_SHEETS_RESULT_HEADERS["comment"]),
        },
    }


def _merged_dict(*items: Any) -> dict[str, Any]:
    merged: dict[str, Any] = {}
    for item in items:
        if isinstance(item, dict):
            merged.update(item)
    return merged


def _result_header(columns: dict[str, Any], *names: str, default: str) -> str:
    for name in names:
        value = columns.get(name)
        if str(value or "").strip():
            return str(value).strip()
    return default


def _positive_int(value: Any, default: int) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed > 0 else default


def _google_sheets_missing_config(config: dict[str, Any]) -> list[str]:
    required = {
        "spreadsheetId": config.get("spreadsheet_id"),
        "sheetName": config.get("sheet_name"),
        "credentialJson": config.get("credential_json"),
    }
    return [name for name, value in required.items() if not str(value or "").strip()]


def _google_sheets_payloads(
    updates: list[dict[str, Any]],
    mapping_entries: list[dict[str, Any]],
    config: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    mapping_by_key = {item.get("automationKey"): item for item in mapping_entries if item.get("automationKey")}
    payloads: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    executed_at = datetime.utcnow().isoformat()
    for item in updates:
        automation_key = item.get("automationKey")
        mapping = mapping_by_key.get(automation_key, {})
        target = _google_sheets_target(mapping)
        spreadsheet_id = str(
            target.get("spreadsheetId")
            or target.get("spreadsheet_id")
            or config.get("spreadsheet_id")
            or ""
        ).strip()
        sheet_name = str(
            target.get("sheetName")
            or target.get("sheet_name")
            or target.get("sheet")
            or config.get("sheet_name")
            or ""
        ).strip()
        row = _target_row(target)
        if not spreadsheet_id:
            issues.append(_issue(
                "google_sheets_spreadsheet_id_missing",
                "Google Sheets spreadsheet ID is required for result write-back",
                automation_key,
            ))
            continue
        if not sheet_name:
            issues.append(_issue(
                "google_sheets_sheet_name_missing",
                "Google Sheets sheet name is required for result write-back",
                automation_key,
            ))
            continue
        if row is None:
            issues.append(_issue(
                "google_sheets_row_missing",
                "Google Sheets row is required for result write-back",
                automation_key,
            ))
            continue
        headers = config.get("headers") if isinstance(config.get("headers"), dict) else GOOGLE_SHEETS_RESULT_HEADERS
        payloads.append({
            "automationKey": automation_key,
            "sourceCaseId": item.get("sourceCaseId"),
            "spreadsheetId": spreadsheet_id,
            "sheetName": sheet_name,
            "row": row,
            "values": {
                headers["status"]: item.get("status") or "",
                headers["run_id"]: item.get("runId") or "",
                headers["executed_at"]: executed_at,
                headers["comment"]: item.get("comment") or "",
            },
        })
    return payloads, issues


def _google_sheets_target(mapping: dict[str, Any]) -> dict[str, Any]:
    targets = mapping.get("resultTargets") or {}
    target = targets.get("googleSheets") or targets.get("google_sheets") or {}
    if target:
        return target if isinstance(target, dict) else {}
    if mapping.get("sourceType") == "google_sheets":
        fallback = targets.get("excel") or {}
        return fallback if isinstance(fallback, dict) else {}
    return {}


def _target_row(target: dict[str, Any]) -> int | None:
    value = target.get("row") or target.get("rowIndex") or target.get("row_index")
    try:
        row = int(value)
    except (TypeError, ValueError):
        return None
    return row if row > 0 else None


async def _post_google_sheets_results(payloads: list[dict[str, Any]], config: dict[str, Any]) -> list[dict[str, Any]]:
    credential_json = str(config.get("credential_json") or "")
    try:
        access_token = await _google_access_token(credential_json, SHEETS_SCOPE)
    except GoogleSheetsConnectorError as exc:
        raise ExportValidationError(_mask_with_credential(exc.message, credential_json)) from exc

    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for payload in payloads:
        grouped.setdefault((payload["spreadsheetId"], payload["sheetName"]), []).append(payload)

    responses: list[dict[str, Any]] = []
    async with httpx.AsyncClient(timeout=30) as client:
        for (spreadsheet_id, sheet_name), group in grouped.items():
            headers = await _fetch_google_sheets_headers(client, spreadsheet_id, sheet_name, access_token, credential_json, config)
            data = _google_sheets_batch_data(sheet_name, headers, group, config)
            if not data:
                continue
            try:
                response = await client.post(
                    _google_sheets_batch_update_url(spreadsheet_id),
                    json={
                        "valueInputOption": str(config.get("value_input_option") or "USER_ENTERED"),
                        "data": data,
                    },
                    headers={
                        "Accept": "application/json",
                        "Authorization": f"Bearer {access_token}",
                    },
                )
            except httpx.HTTPError as error:
                raise ExportValidationError(
                    _mask_with_credential(f"Google Sheets API request failed: {error}", credential_json)
                ) from error
            _raise_google_sheets_response_error(response, credential_json)
            responses.append({
                "spreadsheetId": spreadsheet_id,
                "sheetName": sheet_name,
                "statusCode": response.status_code,
                "updatedRows": len(group),
                "response": _safe_google_response_json(response, credential_json),
            })
    return responses


async def _fetch_google_sheets_headers(
    client: httpx.AsyncClient,
    spreadsheet_id: str,
    sheet_name: str,
    access_token: str,
    credential_json: str,
    config: dict[str, Any],
) -> list[str]:
    header_row = int(config.get("header_row") or 1)
    try:
        response = await client.get(
            _google_sheets_values_url(spreadsheet_id, f"{_a1_sheet_name(sheet_name)}!{header_row}:{header_row}"),
            headers={
                "Accept": "application/json",
                "Authorization": f"Bearer {access_token}",
            },
        )
    except httpx.HTTPError as error:
        raise ExportValidationError(
            _mask_with_credential(f"Google Sheets API request failed: {error}", credential_json)
        ) from error
    _raise_google_sheets_response_error(response, credential_json)
    payload = _safe_response_json(response)
    values = payload.get("values", []) if isinstance(payload, dict) else []
    first = values[0] if values and isinstance(values[0], list) else []
    return [str(value or "").strip() for value in first]


def _google_sheets_batch_data(
    sheet_name: str,
    headers: list[str],
    payloads: list[dict[str, Any]],
    config: dict[str, Any],
) -> list[dict[str, Any]]:
    header_row = int(config.get("header_row") or 1)
    data: list[dict[str, Any]] = []
    header_columns = {header: index for index, header in enumerate(headers, start=1) if header}
    result_headers = list((config.get("headers") or GOOGLE_SHEETS_RESULT_HEADERS).values())
    for header in result_headers:
        if header in header_columns:
            continue
        headers.append(header)
        column = len(headers)
        header_columns[header] = column
        data.append({
            "range": _google_sheets_cell_range(sheet_name, column, header_row),
            "values": [[header]],
        })

    for payload in payloads:
        row = int(payload["row"])
        values = payload.get("values") if isinstance(payload.get("values"), dict) else {}
        for header, value in values.items():
            column = header_columns.get(header)
            if not column:
                continue
            data.append({
                "range": _google_sheets_cell_range(sheet_name, column, row),
                "values": [[value]],
            })
    return data


def _google_sheets_cell_range(sheet_name: str, column: int, row: int) -> str:
    cell = f"{_column_name(column)}{row}"
    return f"{_a1_sheet_name(sheet_name)}!{cell}:{cell}"


def _a1_sheet_name(sheet_name: str) -> str:
    return "'" + sheet_name.replace("'", "''") + "'"


def _column_name(index: int) -> str:
    name = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name or "A"


def _google_sheets_values_url(spreadsheet_id: str, range_name: str) -> str:
    return (
        f"https://sheets.googleapis.com/v4/spreadsheets/{quote(spreadsheet_id)}/values/"
        f"{quote(range_name, safe='!:')}?majorDimension=ROWS"
    )


def _google_sheets_batch_update_url(spreadsheet_id: str) -> str:
    return f"https://sheets.googleapis.com/v4/spreadsheets/{quote(spreadsheet_id)}/values:batchUpdate"


def _raise_google_sheets_response_error(response: httpx.Response, credential_json: str) -> None:
    if response.status_code < 400:
        return
    detail = _google_sheets_error_detail(response, credential_json)
    if response.status_code in {401, 403}:
        raise ExportValidationError(f"Google Sheets credentials rejected or unauthorized. {detail}")
    if response.status_code == 404:
        raise ExportValidationError(f"Google spreadsheet or sheet range was not found. {detail}")
    raise ExportValidationError(f"Google Sheets API returned HTTP {response.status_code}. {detail}")


def _google_sheets_error_detail(response: httpx.Response, credential_json: str) -> str:
    try:
        payload = response.json()
    except ValueError:
        payload = response.text
    if isinstance(payload, dict):
        error = payload.get("error")
        if isinstance(error, dict):
            detail = error.get("message") or str(error)
        else:
            detail = payload.get("message") or str(payload)
    else:
        detail = str(payload)
    return _mask_with_credential(detail, credential_json)


def _safe_google_response_json(response: httpx.Response, credential_json: str) -> Any:
    return _mask_google_secret_data(_safe_response_json(response), credential_json)


def _mask_google_secret_data(value: Any, credential_json: str) -> Any:
    if isinstance(value, str):
        return _mask_with_credential(value, credential_json)
    if isinstance(value, list):
        return [_mask_google_secret_data(item, credential_json) for item in value]
    if isinstance(value, dict):
        return {key: _mask_google_secret_data(item, credential_json) for key, item in value.items()}
    return value


async def export_testrail_clone(session: Session, execution: ExecutionRun, preview: bool = False) -> dict:
    settings = load_settings()
    base_url = settings.integrations.get("testrailClone", {}).get("baseUrl", "http://localhost:3000")
    updates = _result_updates(execution)
    mapping_entries, mapping_issues = _load_mapping_entries(execution)
    validation = _validate_export(session, execution, updates, mapping_entries, mapping_issues)
    payload = {
        "runId": execution.run_id,
        "results": [
            {
                "automationKey": item.get("automationKey"),
                "sourceType": item.get("sourceType"),
                "sourceCaseId": item.get("sourceCaseId"),
                "title": item.get("title"),
                "status": item.get("status"),
                "durationMs": item.get("durationMs"),
                "comment": item.get("comment"),
                "artifacts": item.get("artifacts"),
            }
            for item in updates
        ],
    }
    if preview:
        return {"preview": True, "payload": payload, "validation": validation}

    _raise_if_invalid(validation)

    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.post(f"{base_url.rstrip('/')}/api/automation/results/bulk", json=payload)
        ok = resp.status_code < 400
        _log_export(
            session,
            execution,
            "testrail-clone",
            "success" if ok else "failed",
            updates,
            {"response": mask_secrets(resp.text[:300])},
        )
        resp.raise_for_status()
        return resp.json()


def export_excel(session: Session, execution: ExecutionRun, preview: bool = False) -> dict:
    results = _load_results(execution)
    result_updates = _result_updates(execution, results)
    mapping_entries, mapping_issues = _load_mapping_entries(execution)
    validation = _validate_export(session, execution, result_updates, mapping_entries, mapping_issues)
    by_key = {c.get("automationKey"): c for c in mapping_entries}

    updates = []
    for case in results.get("cases", []):
        key = case.get("automationKey")
        meta = by_key.get(key, {})
        excel_meta = (meta.get("resultTargets") or {}).get("excel") or {}
        if not excel_meta.get("file"):
            continue
        updates.append({
            "automationKey": key,
            "sourceCaseId": case.get("sourceCaseId"),
            "file": excel_meta.get("file"),
            "row": excel_meta.get("row"),
            "status": case.get("status"),
            "runId": execution.run_id,
            "comment": case.get("error") or "Automation passed",
        })

    if preview:
        return {"preview": True, "updates": updates, "validation": validation}

    _raise_if_invalid(validation)

    failed = []
    for item in updates:
        src = Path(item["file"])
        if not src.exists():
            failed.append({**item, "error": "source file not found"})
            continue
        backup = src.with_suffix(src.suffix + f".backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}")
        shutil.copy2(src, backup)
        wb = load_workbook(src)
        ws = wb.active
        row = item["row"]
        headers = [cell.value for cell in ws[1]]
        col_map = {
            "Automation Result": item["status"],
            "Automation Run ID": execution.run_id,
            "Automation Executed At": datetime.utcnow().isoformat(),
            "Automation Comment": item.get("comment") or "",
        }
        for col_name, value in col_map.items():
            if col_name in headers:
                column = headers.index(col_name) + 1
            else:
                column = len(headers) + 1
                ws.cell(row=1, column=column, value=col_name)
                headers.append(col_name)
            ws.cell(row=row, column=column, value=value)
        wb.save(src)

    updated = len(updates) - len(failed)
    _log_export(
        session,
        execution,
        "excel",
        "failed" if failed else "success",
        updates,
        {"failed": failed},
    )
    return {"updated": updated, "failed": failed}


async def export_testrail(
    session: Session,
    execution: ExecutionRun,
    preview: bool = False,
    config: dict[str, Any] | None = None,
) -> dict:
    updates = _result_updates(execution)
    mapping_entries, mapping_issues = _load_mapping_entries(execution)
    validation = _validate_export(session, execution, updates, mapping_entries, mapping_issues)
    export_config = _testrail_export_config(config)
    payloads: list[dict[str, Any]] = []
    if not export_config["mock"]:
        payloads, target_issues = _testrail_payloads(updates, mapping_entries, export_config)
        validation = _merge_validation_issues(validation, target_issues)
    if preview:
        response = {"preview": True, "updates": updates, "validation": validation}
        if payloads:
            response["targetPayload"] = payloads
        return response
    _raise_if_invalid(validation)
    if export_config["mock"]:
        _log_export(session, execution, "testrail", "success", updates, {"mode": "local-mock"})
        return {"updated": len(updates), "target": "testrail", "updates": updates, "mode": "local-mock"}

    missing = _testrail_missing_config(export_config)
    if missing:
        raise ExportValidationError(
            f"TestRail export requires {', '.join(missing)}. Configure TestRail in Settings and store the API token."
        )

    try:
        responses = await _post_testrail_results(payloads, export_config)
    except ExportValidationError as exc:
        _log_export(
            session,
            execution,
            "testrail",
            "failed",
            updates,
            {"error": _mask_with_token(str(exc), str(export_config.get("api_token") or ""))},
        )
        raise

    _log_export(
        session,
        execution,
        "testrail",
        "success",
        updates,
        {"mode": "api", "responses": responses},
    )
    return {"updated": len(responses), "target": "testrail", "updates": updates, "responses": responses}


async def export_google_sheets(
    session: Session,
    execution: ExecutionRun,
    preview: bool = False,
    config: dict[str, Any] | None = None,
) -> dict:
    updates = _result_updates(execution)
    mapping_entries, mapping_issues = _load_mapping_entries(execution)
    validation = _validate_export(session, execution, updates, mapping_entries, mapping_issues)
    export_config = _google_sheets_export_config(config)
    payloads: list[dict[str, Any]] = []
    if not export_config["mock"]:
        payloads, target_issues = _google_sheets_payloads(updates, mapping_entries, export_config)
        validation = _merge_validation_issues(validation, target_issues)
    if preview:
        response = {"preview": True, "updates": updates, "validation": validation}
        if payloads:
            response["targetPayload"] = payloads
        return response
    _raise_if_invalid(validation)
    if export_config["mock"]:
        _log_export(session, execution, "google-sheets", "success", updates, {"mode": "local-mock"})
        return {"updated": len(updates), "target": "google-sheets", "updates": updates, "mode": "local-mock"}

    missing = _google_sheets_missing_config(export_config)
    if missing:
        raise ExportValidationError(
            f"Google Sheets export requires {', '.join(missing)}. Configure Google Sheets in Settings and store credential JSON."
        )

    try:
        responses = await _post_google_sheets_results(payloads, export_config)
    except ExportValidationError as exc:
        _log_export(
            session,
            execution,
            "google-sheets",
            "failed",
            updates,
            {"error": _mask_with_credential(str(exc), str(export_config.get("credential_json") or ""))},
        )
        raise

    _log_export(
        session,
        execution,
        "google-sheets",
        "success",
        updates,
        {"mode": "api", "responses": responses},
    )
    return {
        "updated": len(payloads),
        "target": "google-sheets",
        "updates": updates,
        "responses": responses,
    }
