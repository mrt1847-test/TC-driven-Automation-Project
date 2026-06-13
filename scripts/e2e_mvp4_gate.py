"""H-04 MVP 4 gate E2E against a live Worker on http://127.0.0.1:8765."""
from __future__ import annotations

import shutil
import sys
import tempfile
import time
from pathlib import Path

import httpx

from e2e_worker_client import worker_client
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
EXCEL = ROOT / "fixtures" / "sample_cases.xlsx"
BASE = "http://127.0.0.1:8765"


def wait_for_first(client: httpx.Client, path: str, timeout_s: float = 20.0) -> dict | None:
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        items = client.get(path).json()
        for item in items:
            if item.get("status") in {"completed", "failed", "cancelled"}:
                return item
        time.sleep(0.25)
    return None


def save_mappings(client: httpx.Client, project_id: str, case_id: str) -> None:
    mappings = client.get(f"/projects/{project_id}/cases/{case_id}/mappings").json()
    reviewed = []
    for index, mapping in enumerate(mappings, start=1):
        reviewed.append({
            **mapping,
            "normalized_step_id": mapping.get("normalized_step_id") or f"flow_{index:03d}",
            "normalized_step_name": f"mvp4_step_{index}",
            "pom_method_name": f"perform_mvp4_step_{index}",
            "status": "mapped",
        })
    client.put(f"/projects/{project_id}/cases/{case_id}/mappings", json={"mappings": reviewed}).raise_for_status()


def write_fake_pytest_fixture(client: httpx.Client, project_id: str) -> None:
    content = "\n".join([
        "import pytest",
        "",
        "def pytest_addoption(parser):",
        "    parser.addoption('--browser', action='store', default='chromium')",
        "    parser.addoption('--headed', action='store', default='false')",
        "",
        "class FakeLocator:",
        "    def click(self):",
        "        return None",
        "    def fill(self, value):",
        "        return None",
        "",
        "class FakePage:",
        "    def goto(self, url):",
        "        return None",
        "    def locator(self, selector):",
        "        return FakeLocator()",
        "    def get_by_role(self, *args, **kwargs):",
        "        return FakeLocator()",
        "",
        "@pytest.fixture",
        "def page():",
        "    return FakePage()",
        "",
    ])
    client.put(
        f"/projects/{project_id}/generated-files/content",
        json={"path": "conftest.py", "content": content},
    ).raise_for_status()


def workbook_value(path: str, row: int, column_name: str):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    return sheet.cell(row=row, column=headers.index(column_name) + 1).value


def main() -> int:
    if not EXCEL.exists():
        print(f"Missing fixture: {EXCEL}", file=sys.stderr)
        return 1

    with tempfile.TemporaryDirectory() as temp_dir:
        source = Path(temp_dir) / "mvp4_cases.xlsx"
        shutil.copy2(EXCEL, source)

        client = worker_client(BASE, timeout=60)
        client.get("/health").raise_for_status()

        project = client.post("/projects", json={"name": "E2E MVP 4 Gate"}).json()
        project_id = project["id"]

        cases = client.post(f"/projects/{project_id}/cases/import/excel", json={"file_path": str(source)}).json()
        case = cases[0]
        case_id = case["id"]
        automation_key = case["automation_key"]
        print("case", automation_key, case["source_id"])

        client.post(f"/projects/{project_id}/webwright-runs", json={"caseIds": [case_id]}).raise_for_status()
        run = wait_for_first(client, f"/projects/{project_id}/webwright-runs")
        if not run or run["status"] != "completed":
            print(f"Webwright run did not complete: {run}", file=sys.stderr)
            return 1

        save_mappings(client, project_id, case_id)
        generated = client.post(f"/projects/{project_id}/generate", json={"caseIds": [case_id]}).json()
        if not Path(generated["generatedProjectPath"]).exists():
            print("Generated path missing", file=sys.stderr)
            return 1
        write_fake_pytest_fixture(client, project_id)

        client.post(
            f"/projects/{project_id}/executions",
            json={
                "env": "stg",
                "browser": "chromium",
                "headed": False,
                "target_type": "case",
                "automation_key": automation_key,
                "result_target": "local",
            },
        ).raise_for_status()
        execution = wait_for_first(client, f"/projects/{project_id}/executions")
        if not execution or not execution.get("result_path"):
            print(f"Execution did not produce results: {execution}", file=sys.stderr)
            return 1

        for target in ["testrail", "google-sheets", "excel"]:
            preview = client.post(
                f"/projects/{project_id}/executions/{execution['id']}/export/{target}",
                json={"preview": True},
            ).json()
            updates = preview["updates"]
            if not updates or updates[0]["sourceCaseId"] != case["source_id"]:
                print(f"{target} preview lost source mapping: {preview}", file=sys.stderr)
                return 1
            exported = client.post(
                f"/projects/{project_id}/executions/{execution['id']}/export/{target}",
                json={"preview": False},
            ).json()
            if exported["updated"] != 1:
                print(f"{target} export failed: {exported}", file=sys.stderr)
                return 1
            print(target, "updated", exported["updated"])

        excel_preview = client.post(
            f"/projects/{project_id}/executions/{execution['id']}/export/excel",
            json={"preview": True},
        ).json()
        excel_update = excel_preview["updates"][0]
        if workbook_value(excel_update["file"], excel_update["row"], "Automation Run ID") != execution["run_id"]:
            print("Excel workbook was not updated", file=sys.stderr)
            return 1

        print("H-04 MVP 4 gate E2E OK")
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
